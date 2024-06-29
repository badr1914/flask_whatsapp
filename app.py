from flask import Flask, request, render_template, redirect, url_for, send_file
import pandas as pd
from openpyxl import load_workbook, Workbook
from twilio.twiml.messaging_response import MessagingResponse
from twilio.rest import Client
import os
import logging
import pdfplumber
import shutil
from datetime import datetime
from PIL import Image
import openai
import re

app = Flask(__name__)

# إعداد Twilio
account_sid = 'AC96c2c9d2afc969c3e07987b5360e5316'
auth_token = 'c64498ee94f1d4a6a495461b9e96ad9e'
client = Client(account_sid, auth_token)

# إعداد OpenAI
openai.api_key = 'sk-proj-b0GwBpOzrf8ABS86gZ8KT3BlbkFJeLZcGGWXMxAkfPK7Pvsj'

# تخزين حالة المستخدم في الذاكرة
user_state = {}

# فئات نوع الشراء
categories = ['الصيانة', 'الأثاث', 'أدوات التنظيف', 'أخرى']

# إعداد التسجيل (logging)
logging.basicConfig(level=logging.INFO)

@app.route('/')
def home():
    return render_template('index.html')

@app.route('/handle_button', methods=['POST'])
def handle_button():
    action = request.form.get('action')
    from_number = request.form.get('From', 'web')
    
    if from_number not in user_state:
        user_state[from_number] = {'state': 'start'}

    state = user_state[from_number]['state']
    if action == 'purchase_order':
        user_state[from_number]['state'] = 'details'
        return render_template('purchase_order.html')
    elif action == 'maintenance_request':
        user_state[from_number]['state'] = 'maintenance_type'
        return render_template('maintenance_request.html')
    elif action == 'contact':
        user_state[from_number]['state'] = 'contact_choice'
        return render_template('contact.html')
    elif action == 'message_manager':
        user_state[from_number]['state'] = 'manager_message'
        return render_template('message_manager.html')
    elif action == 'admin':
        user_state[from_number]['state'] = 'admin_login'
        return render_template('admin_login.html')
    elif action == 'request_report':
        user_state[from_number]['state'] = 'report_request'
        return render_template('report_request.html')
    else:
        return redirect(url_for('home'))

@app.route('/handle_purchase_order', methods=['POST'])
def handle_purchase_order():
    details = request.form.get('details')
    amount = request.form.get('amount')
    purchase_type = request.form.get('type')
    from_number = request.form.get('From', 'web')

    if from_number in user_state:
        user_state[from_number]['details'] = details
        user_state[from_number]['amount'] = amount
        user_state[from_number]['type'] = purchase_type

        save_to_invoices_excel(user_state[from_number], "dummy_path.pdf")
        del user_state[from_number]
        return "تم حفظ أمر الشراء بنجاح!"

@app.route('/upload_pdf', methods=['GET', 'POST'])
def upload_pdf():
    if request.method == 'POST':
        if 'pdf' not in request.files:
            return "لم يتم تحميل أي ملف PDF", 400

        from_number = request.form.get('From')
        if not from_number:
            return "رقم المرسل غير موجود في الطلب.", 400

        if from_number not in user_state or user_state[from_number]['state'] != 'awaiting_pdf':
            return "لم يتم طلب ملف PDF حاليا.", 400

        pdf_file = request.files['pdf']
        if not os.path.exists('uploads'):
            os.makedirs('uploads')
        pdf_path = os.path.join('uploads', pdf_file.filename)
        pdf_file.save(pdf_path)

        amount = extract_amount_from_pdf(pdf_path)
        input_amount = user_state[from_number]['amount']
        if amount != input_amount:
            return f"المبلغ المستخرج من ملف PDF لا يطابق المبلغ المدخل: {amount} != {input_amount}", 400

        save_to_invoices_excel(user_state[from_number], pdf_path)

        # نقل ملف PDF إلى مجلد محدد
        target_dir = 'processed_pdfs'
        if not os.path.exists(target_dir):
            os.makedirs(target_dir)
        shutil.move(pdf_path, os.path.join(target_dir, pdf_file.filename))

        del user_state[from_number]
        return "تم تحميل الملف بنجاح ومعالجة المبالغ."
    
    from_number = request.args.get('From')
    if not from_number:
        return "رقم المرسل غير موجود.", 400
    
    return '''
    <!doctype html>
    <title>Upload PDF</title>
    <h1>Upload PDF</h1>
    <form method=post enctype=multipart/form-data>
      <input type=file name=pdf>
      <input type=hidden name=From value="{}">
      <input type=submit value=Upload>
    </form>
    '''.format(from_number)

def extract_amount_from_text(text):
    logging.info(f"Extracted text: {text[:1000]}...")  # عرض أول 1000 حرف فقط للتأكد من القراءة
    # نمط لاستخراج المبلغ الذي يحتوي على الفاصلة والرقم العشري
    match = re.search(r'المبلغ\s*:\s*(\d+)', text)
    if match:
        amount = match.group(1).replace(',', '')
        return int(amount)  # تحويل المبلغ إلى رقم صحيح
    else:
        raise ValueError("لم يتم العثور على المبلغ في ملف PDF")

def extract_amount_from_pdf(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ''
        for page in pdf.pages:
            text += page.extract_text()

    logging.info(f"Full extracted text: {text[:1000]}...")  # تسجيل أول 1000 حرف فقط
    # استخدم الدالة لاستخراج المبلغ من النص
    return extract_amount_from_text(text)

def save_to_invoices_excel(purchase_data, pdf_path):
    file_path = 'invoices.xlsx'
    sheet_name = 'invoices'

    if not os.path.exists(file_path):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        headers = ['تاريخ الشراء', 'المبلغ', 'التفاصيل', 'نوع الشراء', 'ملف PDF']
        ws.append(headers)
        wb.save(file_path)
    
    book = load_workbook(file_path)
    if sheet_name not in book.sheetnames:
        sheet = book.create_sheet(title=sheet_name)
        sheet.append(['تاريخ الشراء', 'المبلغ', 'التفاصيل', 'نوع الشراء', 'ملف PDF'])
    else:
        sheet = book[sheet_name]
    
    data = [[datetime.now().strftime("%Y-%m-%d"), purchase_data['amount'], purchase_data['details'], purchase_data['type'], os.path.basename(pdf_path)]]
    
    for row in data:
        sheet.append(row)
    
    book.save(file_path)

# Other helper functions ...

if __name__ == '__main__':
    app.run(debug=True, port=5001)
